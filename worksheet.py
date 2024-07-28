import openpyxl
import os


def generate(data):
    if os.path.exists('Acompanhamento de preços.xlsx'): 
        workbook = openpyxl.load_workbook('Acompanhamento de preços.xlsx')
    else:
        workbook = openpyxl.Workbook()
        del workbook['Sheet']
        workbook.create_sheet('precos')
        sheet_vagas = workbook['precos']
        sheet_vagas.append(
            ['Data', 'Produto', 'Data da consulta', 'Link'])
        
    sheet_vagas = workbook['precos']
    sheet_vagas.append(data)

    workbook.save('Acompanhamento de preços.xlsx')